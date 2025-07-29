import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Set, Any
from pathlib import Path

class ScheduleClass:
    """Class representing a scheduled lesson with all its properties."""
    
    def __init__(self, 
                 subject: str, 
                 group: str, 
                 teacher: str, 
                 main_room: str,
                 alternative_rooms: List[str],
                 building: str,
                 duration: int,
                 day: str,
                 start_time: Optional[str] = None,
                 end_time: Optional[str] = None,
                 pause_before: Optional[int] = None,
                 pause_after: Optional[int] = None,
                 section_index: int = 0,
                 column: str = "B"):
        
        self.subject = subject
        self.group = group
        self.teacher = teacher
        self.main_room = main_room
        self.alternative_rooms = [r for r in alternative_rooms if r]  # Filter out None values
        self.building = building
        self.duration = int(duration) if duration is not None else 0
        self.day = day
        self.start_time = start_time
        self.end_time = end_time
        self.pause_before = int(pause_before) if pause_before is not None else 0
        self.pause_after = int(pause_after) if pause_after is not None else 0
        self.section_index = section_index
        self.column = column
        
        # Добавляем новые атрибуты для работы с временными окнами
        self.has_time_window = False  # Этот флаг будет установлен в model_variables.py
        self.fixed_start_time = start_time is not None and end_time is None
        
        # Linked classes (to be filled later)
        self.next_class = None
        self.previous_class = None
        self.linked_classes = []
        
    def __str__(self):
        time_info = "No time"
        if self.start_time and self.end_time:
            time_info = f"{self.start_time}-{self.end_time}"
        elif self.start_time:
            time_info = self.start_time
            
        return f"{self.subject} - {self.group} - {self.teacher} - {self.day} {time_info}"
    
    def __repr__(self):
        return self.__str__()
    
    @property
    def possible_rooms(self) -> List[str]:
        """Return the list of all possible rooms for this class."""
        rooms = [self.main_room] + self.alternative_rooms
        return [r for r in rooms if r]  # Filter out any None values
    
    @property
    def has_fixed_time(self) -> bool:
        """Check if the class has a fixed start time."""
        return self.start_time is not None and self.end_time is None
    
    @property
    def has_fixed_room(self) -> bool:
        """Check if the class must be in a specific room."""
        return len(self.alternative_rooms) == 0
    
    @property
    def total_duration(self) -> int:
        """Calculate the total duration including pauses."""
        return self.duration + self.pause_before + self.pause_after
    
    def get_groups(self) -> List[str]:
        """Extract all group names from the group field."""
        # The group field may contain multiple groups like "2A+1A Kunst"
        # Parse this to extract all group names
        if not self.group:
            return []
            
        parts = self.group.split()
        groups = []
        
        for part in parts:
            # Look for patterns like "2A", "1A+3B", etc.
            if any(c.isdigit() for c in part):
                if "+" in part:
                    # Handle multiple groups
                    for group_part in part.split("+"):
                        groups.append(group_part.strip())
                else:
                    groups.append(part)
                    
        return groups if groups else [self.group]


class ScheduleReader:
    """Class for reading and processing schedule data from Excel files."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.planning_sections = []
        self.teachers = set()
        self.groups = set()
        self.rooms = set()
        self.buildings = set()
        self.days = set()
        
    def read_excel(self) -> List[ScheduleClass]:
        """Read and parse the Excel file to extract scheduling data."""
        # Read the Excel file with both data_only and with formulas
        workbook_data = openpyxl.load_workbook(self.file_path, data_only=True)
        
        # Get the planning sheet
        planning_sheet = None
        for sheet_name in workbook_data.sheetnames:
            if sheet_name.lower() == "plannung":
                planning_sheet = workbook_data[sheet_name]
                break
        
        if not planning_sheet:
            raise ValueError("Could not find 'Plannung' sheet in the Excel file")
        
        # Helper function to extract time values from cells
        def extract_time(row, column):
            # Try different methods to get the time
            cell = planning_sheet.cell(row=row, column=column)
            cell_value = cell.value
            
            # If it's a datetime.time object
            if hasattr(cell_value, 'hour') and hasattr(cell_value, 'minute'):
                return f"{cell_value.hour:02d}:{cell_value.minute:02d}"
                
            # If it's already a datetime, format it
            if isinstance(cell_value, datetime):
                return cell_value.strftime('%H:%M')
                
            # If it's a number, treat it as Excel time (fraction of day)
            if isinstance(cell_value, (int, float)):
                # Convert Excel time to hours and minutes
                total_minutes = int(cell_value * 24 * 60)
                hours = total_minutes // 60
                minutes = total_minutes % 60
                return f"{hours:02d}:{minutes:02d}"
            
            # For string values that look like time (HH:MM:SS)
            if isinstance(cell_value, str) and ':' in cell_value:
                parts = cell_value.split(':')
                if len(parts) >= 2:
                    try:
                        hours = int(parts[0])
                        minutes = int(parts[1])
                        return f"{hours:02d}:{minutes:02d}"
                    except ValueError:
                        pass
            
            # For other values or None
            return cell_value
        
        # Extract planning sections
        planning_map = {}
        row = 2  # Start from row 2
        
        while row <= planning_sheet.max_row:
            # Check if this is the start of a section
            cell_value = planning_sheet.cell(row=row, column=2).value  # Column B
            
            if cell_value:  # Found a section
                section_index = (row - 2) // 14
                planning_map[section_index] = {}
                
                # Process columns B, C, D
                for col_idx, col_letter in enumerate(['B', 'C', 'D'], start=2):
                    subject = planning_sheet.cell(row=row, column=col_idx).value
                    
                    if not subject:
                        continue  # Skip empty columns
                    
                    # Extract all data for this class
                    group = planning_sheet.cell(row=row+1, column=col_idx).value
                    teacher = planning_sheet.cell(row=row+2, column=col_idx).value
                    main_room = planning_sheet.cell(row=row+3, column=col_idx).value
                    alt_room1 = planning_sheet.cell(row=row+4, column=col_idx).value
                    alt_room2 = planning_sheet.cell(row=row+5, column=col_idx).value
                    alt_room3 = planning_sheet.cell(row=row+6, column=col_idx).value
                    building = planning_sheet.cell(row=row+7, column=col_idx).value
                    duration = planning_sheet.cell(row=row+8, column=col_idx).value
                    day = planning_sheet.cell(row=row+9, column=col_idx).value
                    
                    # Use our custom function to extract time values
                    start_time = extract_time(row+10, col_idx)  # Start time cell
                    end_time = extract_time(row+11, col_idx)    # End time cell
                    
                    pause_before = planning_sheet.cell(row=row+12, column=col_idx).value
                    pause_after = planning_sheet.cell(row=row+13, column=col_idx).value
                    
                    # Ensure numeric values for duration and pauses
                    try:
                        duration = int(float(duration)) if duration is not None else 0
                    except (ValueError, TypeError):
                        print(f"Warning: Invalid duration value '{duration}' for {subject}. Using 0.")
                        duration = 0
                        
                    try:
                        pause_before = int(float(pause_before)) if pause_before is not None else 0
                    except (ValueError, TypeError):
                        pause_before = 0
                        
                    try:
                        pause_after = int(float(pause_after)) if pause_after is not None else 0
                    except (ValueError, TypeError):
                        pause_after = 0
                    
                    # Create ScheduleClass object
                    class_data = ScheduleClass(
                        subject=subject,
                        group=group,
                        teacher=teacher,
                        main_room=main_room,
                        alternative_rooms=[alt_room1, alt_room2, alt_room3],
                        building=building,
                        duration=duration,
                        day=day,
                        start_time=start_time,
                        end_time=end_time,
                        pause_before=pause_before,
                        pause_after=pause_after,
                        section_index=section_index,
                        column=col_letter
                    )
                    
                    # Update sets of teachers, groups, rooms, buildings, days
                    self.teachers.add(teacher)
                    self.buildings.add(building) if building else None
                    self.days.add(day) if day else None
                    
                    if main_room:
                        self.rooms.add(main_room)
                    
                    for room in [alt_room1, alt_room2, alt_room3]:
                        if room:
                            self.rooms.add(room)
                    
                    for group_name in class_data.get_groups():
                        self.groups.add(group_name)
                    
                    planning_map[section_index][col_letter] = class_data
                
                # Skip to the next section
                row += 14
            else:
                row += 1
        
        # Connect linked classes
        for section_idx, section in planning_map.items():
            if 'B' in section:
                main_class = section['B']
                main_class.linked_classes = []  # Initialize empty list
                
                if 'C' in section:
                    # ИСПРАВЛЕНО: previous_class теперь ссылка на объект, а не строка
                    section['C'].previous_class = main_class
                    main_class.next_class = section['C']
                    main_class.linked_classes.append(section['C'])
                    
                    if 'D' in section:
                        # ИСПРАВЛЕНО: previous_class теперь ссылка на объект, а не строка
                        section['D'].previous_class = section['C']
                        section['C'].next_class = section['D']
                        main_class.linked_classes.append(section['D'])
        
        # Collect all classes including linked ones
        all_classes = []
        for section_idx, section in planning_map.items():
            for col in ['B', 'C', 'D']:
                if col in section:
                    all_classes.append(section[col])
        
        # Debugging output to check all classes
        print(f"\nAll classes to be processed:")
        for i, cls in enumerate(all_classes):
            print(f"  {i}: {cls}")
        
        # Check for linked classes
        for cls in all_classes:
            if cls.linked_classes:
                print(f"\nClass {cls} has linked classes:")
                for linked in cls.linked_classes:
                    print(f"  - {linked}")
                    # Verify linked class is in the all_classes list
                    if linked not in all_classes:
                        print(f"    WARNING: This linked class is not in the all_classes list!")
        
        self.planning_sections = all_classes
        return all_classes
    
    def _format_time(self, time_value: Any) -> Optional[str]:
        """Format Excel time values to HH:MM string format."""
        try:
            if time_value is None:
                return None
                
            if isinstance(time_value, str):
                return time_value
                
            if isinstance(time_value, datetime):
                return time_value.strftime('%H:%M')
            
            # Handle datetime.time objects
            if hasattr(time_value, 'hour') and hasattr(time_value, 'minute'):
                return f"{time_value.hour:02d}:{time_value.minute:02d}"
                
            # Handle Excel time as decimal fraction of day
            if isinstance(time_value, (int, float)):
                # Convert decimal day fraction to minutes
                minutes_total = int(time_value * 24 * 60)
                hours = minutes_total // 60
                minutes = minutes_total % 60
                return f"{hours:02d}:{minutes:02d}"
                
            # Try to convert other types
            return str(time_value)
        except Exception as e:
            print(f"Warning: Error formatting time value '{time_value}': {str(e)}")
            return None
    
    def get_time_slots(self, interval_minutes=15) -> List[str]:
        """Generate time slots for scheduling at given intervals."""
        # Standard school day from 8:00 to 20:00
        time_slots = []
        current_time = datetime.strptime("08:00", "%H:%M")
        end_time = datetime.strptime("20:00", "%H:%M")
        
        while current_time <= end_time:
            time_slots.append(current_time.strftime("%H:%M"))
            current_time += timedelta(minutes=interval_minutes)
            
        return time_slots
    
    def get_day_indices(self) -> Dict[str, int]:
        """Map day abbreviations to indices (Mo=0, Di=1, etc.)."""
        day_order = ["Mo", "Di", "Mi", "Do", "Fr", "Sa"]
        return {day: idx for idx, day in enumerate(day_order)}


def main():
    """Example usage of the ScheduleReader class."""
    file_path = Path("xlsx_initial") / "schedule_planning.xlsx"
    reader = ScheduleReader(file_path)
    
    try:
        classes = reader.read_excel()
        
        print(f"Successfully read {len(classes)} classes from the Excel file.")
        print(f"Teachers: {reader.teachers}")
        print(f"Groups: {reader.groups}")
        print(f"Rooms: {reader.rooms}")
        print(f"Buildings: {reader.buildings}")
        print(f"Days: {reader.days}")
        
        # Print some sample classes
        print("\nSample Classes:")
        for i, class_data in enumerate(classes[:5]):  # Print first 5 classes
            print(f"\nClass {i+1}: {class_data}")
            print(f"  Subject: {class_data.subject}")
            print(f"  Teacher: {class_data.teacher}")
            print(f"  Group: {class_data.group}")
            print(f"  Duration: {class_data.duration} minutes")
            print(f"  Day: {class_data.day}")
            print(f"  Start Time (raw): {class_data.start_time} (type: {type(class_data.start_time).__name__})")
            print(f"  End Time (raw): {class_data.end_time} (type: {type(class_data.end_time).__name__})")
            print(f"  Possible Rooms: {class_data.possible_rooms}")
            
            if class_data.linked_classes:
                print(f"  Linked Classes: {[c.subject for c in class_data.linked_classes]}")
    
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()